"""研发项目收益的源数据加载、计算、校验和可选落库 Tasks。"""

import hashlib
import os
import re
from datetime import date
from pathlib import Path
from typing import Dict, Mapping, Optional, Sequence, Tuple
from urllib.parse import quote

import numpy as np
import pandas as pd
from mypackage.utilities import connect_to_db

from prefect import task

UNMAPPED_PRODUCT = "未映射产品"
TARGET_TABLE_PATTERN = re.compile(r"^[a-z_][a-z0-9_]*$")
EXPENSE_BASE_BUSINESS_LINE = "国际业务"
RD_REVENUE_SECONDARY_ORG = "国际业务中心"
DEFAULT_OUTPUT_DIR = "/root/prefect/check/rd_project_profitability"

RESULT_COLUMNS = [
    "period_start",
    "period_end",
    "product_sub_category",
    "project_stage",
    "mapping_status",
    "sales_quantity",
    "rd_related_revenue",
    "rd_related_cost",
    "rd_related_gross_profit",
    "rd_related_gross_margin",
    "total_expense",
    "remaining_profit",
    "oa_expense",
    "material_expense",
    "mold_expense",
    "allocated_rd_expense",
    "tech_maintenance_expense",
    "allocated_sales_expense",
    "allocated_management_expense",
    "income_share",
    "labor_share",
    "cost_share",
    "rd_expense_base",
    "management_expense_base",
    "sales_expense_base",
    "rd_allocatable_pool",
    "sales_allocatable_pool",
    "source_no",
    "calculated_at",
]

REVENUE_COST_BACKUP_COLUMNS = [
    "period_start",
    "period_end",
    "product_sub_category",
    "mapped_product_sub_category",
    "mapping_status",
    "material_code",
    "revenue_category",
    "primary_org",
    "secondary_org",
    "revenue",
    "cost",
    "gross_profit",
]


def get_default_rd_project_period(
    reference_date: Optional[date] = None,
) -> Tuple[pd.Timestamp, pd.Timestamp]:
    """返回参考日期之前最近已结束月份的年初至月末期间。"""
    reference = pd.Timestamp(reference_date or date.today()).normalize()
    current_month_start = reference.replace(day=1)
    end_date = current_month_start - pd.Timedelta(days=1)
    start_date = pd.Timestamp(year=end_date.year, month=1, day=1)
    return start_date, end_date


def resolve_rd_project_period(
    start_date: Optional[str],
    end_date: Optional[str],
    reference_date: Optional[date] = None,
) -> Tuple[pd.Timestamp, pd.Timestamp]:
    """解析显式日期；两者都省略时使用最近已结月的年初至今。"""
    if start_date is None and end_date is None:
        return get_default_rd_project_period(reference_date)
    if start_date is None or end_date is None:
        raise ValueError("start_date 和 end_date 必须同时提供，或同时留空使用默认期间")

    try:
        period_start = pd.Timestamp(start_date).normalize()
        period_end = pd.Timestamp(end_date).normalize()
    except (TypeError, ValueError) as exc:
        raise ValueError("start_date 和 end_date 必须是有效日期") from exc

    if period_start > period_end:
        raise ValueError("start_date 不能晚于 end_date")
    if period_start.year < 2020:
        raise ValueError("研发项目收益期间不得早于 2020-01-01")
    return period_start, period_end


def _fetch_dataframe(cursor, sql: str, params: Sequence[object]) -> pd.DataFrame:
    cursor.execute(sql, params)
    return pd.DataFrame(
        cursor.fetchall(), columns=[description[0] for description in cursor.description]
    )


@task(name="load_rd_project_profitability_sources", log_prints=True)
def load_rd_project_profitability_sources_task(
    period_start: pd.Timestamp,
    period_end: pd.Timestamp,
) -> Dict[str, pd.DataFrame]:
    """按指定期间读取 Power BI 研发分析所需的 PostgreSQL 源数据。"""
    conn, cur = connect_to_db()
    date_params = (period_start.date(), period_end.date())
    try:
        sources = {
            "project_stages": _fetch_dataframe(
                cur,
                """
                SELECT product_sub_category, proj_status
                FROM excel_dev_projects
                WHERE date >= %s AND date <= %s
                """,
                date_params,
            ),
            "material_master": _fetch_dataframe(
                cur,
                """
                SELECT encoding, name, material_group, material_major_category
                FROM dim_material_master
                WHERE material_group NOT IN ('中正成品', '研发手装样机成品')
                  AND COALESCE(material_major_category, material_group) IN ('产成品', '服务')
                """,
                (),
            ),
            "product_overrides": _fetch_dataframe(
                cur,
                """
                SELECT encoding, product_sub_category
                FROM excel_sub_prod
                """,
                (),
            ),
            "revenue": _fetch_dataframe(
                cur,
                """
                SELECT
                    r.mat_code,
                    r.inc_major_cat,
                    o.prim_org,
                    o.sec_org,
                    SUM(
                        CASE
                            WHEN COALESCE(r.amt_tax_exc_loc, 0) <> 0
                            THEN COALESCE(r.sales_qty, 0)
                            ELSE 0
                        END
                    ) AS sales_quantity,
                    SUM(COALESCE(r.amt_tax_exc_loc, 0)) AS revenue,
                    SUM(
                        COALESCE(r.cost_amt, 0)
                        + COALESCE(r.freight_cost, 0)
                        + COALESCE(r.soft_cost, 0)
                        + COALESCE(r.tariff_cost, 0)
                    ) AS cost
                FROM fact_revenue r
                LEFT JOIN dim_org_struc o ON o.unique_lvl = r.unique_lvl
                WHERE r.acct_period >= %s
                  AND r.acct_period <= %s
                  AND o.prim_org_map IS DISTINCT FROM '无归属'
                GROUP BY r.mat_code, r.inc_major_cat, o.prim_org, o.sec_org
                """,
                date_params,
            ),
            "oa_expenses": _fetch_dataframe(
                cur,
                """
                SELECT e.product_sub_category, SUM(e.exp_amt) AS amount
                FROM fact_expense e
                JOIN dim_org_struc o ON o.unique_lvl = e.unique_lvl
                WHERE e.acct_period >= %s
                  AND e.acct_period <= %s
                  AND o.map_rel IN ('技术中心', '产品中心')
                  AND o.prim_org_map = '支付硬件事业群'
                  AND e.exp_nature = '变动费用'
                  AND COALESCE(e.summary, '') NOT LIKE '%%研发费用%%'
                  AND COALESCE(e.summary, '') NOT LIKE '%%其他出库%%'
                  AND COALESCE(e.summary, '') NOT LIKE '%%其他入库%%'
                GROUP BY e.product_sub_category
                """,
                date_params,
            ),
            "labor_hours": _fetch_dataframe(
                cur,
                """
                SELECT product_sub_category, SUM(hours_worked) AS amount
                FROM excel_labor_hours
                WHERE date >= %s AND date <= %s
                GROUP BY product_sub_category
                """,
                date_params,
            ),
            "material_expenses": _fetch_dataframe(
                cur,
                """
                SELECT product_sub_category, SUM(accounted_amount) AS amount
                FROM excel_material_usage_costs
                WHERE date >= %s AND date <= %s
                GROUP BY product_sub_category
                """,
                date_params,
            ),
            "mold_expenses": _fetch_dataframe(
                cur,
                """
                SELECT product_sub_category, SUM(amt_tax_exc_loc) AS amount
                FROM excel_sample_molds
                WHERE bus_date >= %s AND bus_date <= %s
                GROUP BY product_sub_category
                """,
                date_params,
            ),
            "tech_maintenance_expenses": _fetch_dataframe(
                cur,
                """
                SELECT product_sub_category, SUM(amt) AS amount
                FROM excel_tech_maintenance
                WHERE date >= %s AND date <= %s
                GROUP BY product_sub_category
                """,
                date_params,
            ),
            "expense_bases": _fetch_dataframe(
                cur,
                """
                SELECT e.prim_subj, SUM(e.exp_amt) AS amount
                FROM fact_bus_expense e
                LEFT JOIN dim_org_struc o ON o.unique_lvl = e.unique_lvl
                WHERE e.acct_period >= %s
                  AND e.acct_period <= %s
                  AND e.prim_subj IN ('研发费用', '管理费用', '销售费用')
                  AND e.bus_line = %s
                  AND o.prim_org_map IS DISTINCT FROM '无归属'
                GROUP BY e.prim_subj
                """,
                date_params + (EXPENSE_BASE_BUSINESS_LINE,),
            ),
        }
    finally:
        cur.close()
        conn.close()

    counts = ", ".join(f"{name}={len(frame)}" for name, frame in sources.items())
    print(f"研发项目收益源数据加载完成：{counts}")
    return sources


def _require_columns(df: pd.DataFrame, required: Sequence[str], source_name: str) -> None:
    missing = sorted(set(required) - set(df.columns))
    if missing:
        raise ValueError(f"{source_name} 缺少必要字段: {', '.join(missing)}")


def _normalize_product(value: object) -> Optional[str]:
    if pd.isna(value):
        return None
    normalized = str(value).strip()
    return normalized or None


def _product_key(value: object) -> Optional[str]:
    """模拟 Power BI 文本关系的大小写不敏感及首尾空格不敏感匹配。"""
    normalized = _normalize_product(value)
    if normalized is None:
        return None
    normalized_spaces = re.sub(r"\s+", " ", normalized.replace("\xa0", " "))
    return normalized_spaces.casefold()


def _derive_product_from_material_name(value: object) -> Optional[str]:
    """复现 Power Query 对物料名称生成产品细类的核心规则。"""
    if pd.isna(value):
        return None
    first_segment = str(value).split("-", 1)[0]
    if first_segment == "技术服务费":
        return "技术服务"
    without_chinese = re.sub(r"[一-龟]", "", first_segment)
    return _normalize_product(without_chinese.split("(", 1)[0])


def _build_product_mapping(
    material_master: pd.DataFrame,
    product_overrides: pd.DataFrame,
) -> Tuple[Dict[str, Optional[str]], set[str]]:
    _require_columns(
        material_master,
        ["encoding", "name", "material_group", "material_major_category"],
        "dim_material_master",
    )
    _require_columns(
        product_overrides,
        ["encoding", "product_sub_category"],
        "excel_sub_prod",
    )

    master = material_master.copy()
    master["encoding"] = master["encoding"].map(_normalize_product)
    master["derived_product"] = master["name"].map(_derive_product_from_material_name)

    overrides = product_overrides.copy()
    overrides["encoding"] = overrides["encoding"].map(_normalize_product)
    overrides["product_sub_category"] = overrides["product_sub_category"].map(_normalize_product)
    override_conflicts = (
        overrides.dropna(subset=["encoding"])
        .groupby("encoding")["product_sub_category"]
        .nunique(dropna=True)
    )
    conflicting_codes = sorted(override_conflicts[override_conflicts > 1].index.tolist())
    if conflicting_codes:
        preview = ", ".join(conflicting_codes[:10])
        raise ValueError(f"excel_sub_prod 同一编码存在多个产品细类: {preview}")

    override_map = (
        overrides.dropna(subset=["encoding"])
        .drop_duplicates("encoding", keep="last")
        .set_index("encoding")["product_sub_category"]
    )
    master["product_sub_category"] = master["encoding"].map(override_map)
    master["product_sub_category"] = master["product_sub_category"].combine_first(
        master["derived_product"]
    )
    master["product_sub_category"] = master["product_sub_category"].map(_normalize_product)

    mapping_conflicts = (
        master.dropna(subset=["encoding"])
        .groupby("encoding")["product_sub_category"]
        .nunique(dropna=True)
    )
    conflicting_master_codes = sorted(mapping_conflicts[mapping_conflicts > 1].index.tolist())
    if conflicting_master_codes:
        preview = ", ".join(conflicting_master_codes[:10])
        raise ValueError(f"产品主数据同一编码映射到多个产品细类: {preview}")

    product_map = (
        master.dropna(subset=["encoding"])
        .drop_duplicates("encoding", keep="last")
        .set_index("encoding")["product_sub_category"]
        .to_dict()
    )
    product_master_values = set(master["product_sub_category"].dropna().map(_product_key).tolist())
    return product_map, product_master_values


def _prepare_project_stages(
    df: pd.DataFrame,
) -> Tuple[list[str], Dict[str, Optional[str]], Dict[str, str]]:
    _require_columns(df, ["product_sub_category", "proj_status"], "excel_dev_projects")
    stages = df.copy()
    stages["product_sub_category"] = stages["product_sub_category"].map(_normalize_product)
    stages = stages.dropna(subset=["product_sub_category"])
    stages["product_key"] = stages["product_sub_category"].map(_product_key)
    selected_product_lookup = (
        stages.drop_duplicates("product_key", keep="first")
        .set_index("product_key")["product_sub_category"]
        .to_dict()
    )
    products = sorted(selected_product_lookup.values(), key=str.casefold)

    stage_map: Dict[str, Optional[str]] = {}
    for product_key, group in stages.groupby("product_key", sort=False):
        values = []
        for value in group["proj_status"].tolist():
            normalized = None if pd.isna(value) else str(value).strip() or None
            if normalized not in values:
                values.append(normalized)
        display_product = selected_product_lookup[product_key]
        stage_map[display_product] = values[0] if len(values) == 1 else None
    return products, stage_map, selected_product_lookup


def _bucket_product(value: object, selected_product_lookup: Mapping[str, str]) -> str:
    product_key = _product_key(value)
    return selected_product_lookup.get(product_key, UNMAPPED_PRODUCT)


def _group_amount_by_bucket(
    df: pd.DataFrame,
    selected_product_lookup: Mapping[str, str],
    source_name: str,
) -> pd.Series:
    _require_columns(df, ["product_sub_category", "amount"], source_name)
    grouped = df.copy()
    grouped["amount"] = pd.to_numeric(grouped["amount"], errors="coerce").fillna(0.0)
    grouped["bucket"] = grouped["product_sub_category"].map(
        lambda value: _bucket_product(value, selected_product_lookup)
    )
    return grouped.groupby("bucket", dropna=False)["amount"].sum()


def _series_to_column(detail: pd.DataFrame, values: pd.Series) -> pd.Series:
    return detail["product_sub_category"].map(values).fillna(0.0).astype(float)


def calculate_rd_project_profitability(
    sources: Mapping[str, pd.DataFrame],
    period_start: pd.Timestamp,
    period_end: pd.Timestamp,
) -> Dict[str, object]:
    """按统一研发项目口径计算产品收益和收入成本备查明细。"""
    required_sources = {
        "project_stages",
        "material_master",
        "product_overrides",
        "revenue",
        "oa_expenses",
        "labor_hours",
        "material_expenses",
        "mold_expenses",
        "tech_maintenance_expenses",
        "expense_bases",
    }
    missing_sources = sorted(required_sources - set(sources))
    if missing_sources:
        raise ValueError(f"研发项目收益缺少源数据: {', '.join(missing_sources)}")

    products, stage_map, selected_product_lookup = _prepare_project_stages(
        sources["project_stages"]
    )
    product_map, product_master_values = _build_product_mapping(
        sources["material_master"], sources["product_overrides"]
    )

    revenue = sources["revenue"].copy()
    _require_columns(
        revenue,
        [
            "mat_code",
            "inc_major_cat",
            "prim_org",
            "sec_org",
            "sales_quantity",
            "revenue",
            "cost",
        ],
        "fact_revenue",
    )
    revenue["mat_code"] = revenue["mat_code"].map(_normalize_product)
    revenue["mapped_product"] = revenue["mat_code"].map(product_map)
    revenue["bucket"] = revenue["mapped_product"].map(
        lambda value: _bucket_product(value, selected_product_lookup)
    )
    for column in ["sales_quantity", "revenue", "cost"]:
        revenue[column] = pd.to_numeric(revenue[column], errors="coerce").fillna(0.0)

    electronic_payment = revenue[revenue["inc_major_cat"].eq("电子支付")]
    technical_service = revenue[revenue["inc_major_cat"].eq("技术服务")]
    rd_income_cost_rows = pd.concat(
        [
            electronic_payment[electronic_payment["sec_org"].eq(RD_REVENUE_SECONDARY_ORG)],
            technical_service[technical_service["sec_org"].eq(RD_REVENUE_SECONDARY_ORG)],
        ],
        ignore_index=True,
    )

    rd_revenue = rd_income_cost_rows.groupby("bucket")["revenue"].sum()
    rd_cost = rd_income_cost_rows.groupby("bucket")["cost"].sum()
    electronic_revenue = electronic_payment.groupby("bucket")["revenue"].sum()
    electronic_cost = electronic_payment.groupby("bucket")["cost"].sum()
    sales_quantity = electronic_payment.groupby("bucket")["sales_quantity"].sum()

    revenue_cost_backup = rd_income_cost_rows[
        [
            "bucket",
            "mapped_product",
            "mat_code",
            "inc_major_cat",
            "prim_org",
            "sec_org",
            "revenue",
            "cost",
        ]
    ].copy()
    revenue_cost_backup = revenue_cost_backup.rename(
        columns={
            "bucket": "product_sub_category",
            "mapped_product": "mapped_product_sub_category",
            "mat_code": "material_code",
            "inc_major_cat": "revenue_category",
            "prim_org": "primary_org",
            "sec_org": "secondary_org",
        }
    )
    revenue_cost_backup["mapping_status"] = np.where(
        revenue_cost_backup["product_sub_category"].eq(UNMAPPED_PRODUCT), "unmapped", "mapped"
    )
    revenue_cost_backup["gross_profit"] = (
        revenue_cost_backup["revenue"] - revenue_cost_backup["cost"]
    )
    revenue_cost_backup["period_start"] = pd.Timestamp(period_start).normalize()
    revenue_cost_backup["period_end"] = pd.Timestamp(period_end).normalize()
    revenue_cost_backup = revenue_cost_backup[REVENUE_COST_BACKUP_COLUMNS].sort_values(
        ["product_sub_category", "revenue_category", "material_code"],
        kind="stable",
        na_position="last",
    )

    labor = _group_amount_by_bucket(
        sources["labor_hours"], selected_product_lookup, "excel_labor_hours"
    )
    material = _group_amount_by_bucket(
        sources["material_expenses"],
        selected_product_lookup,
        "excel_material_usage_costs",
    )
    mold = _group_amount_by_bucket(
        sources["mold_expenses"], selected_product_lookup, "excel_sample_molds"
    )
    tech = _group_amount_by_bucket(
        sources["tech_maintenance_expenses"],
        selected_product_lookup,
        "excel_tech_maintenance",
    )

    oa_source = sources["oa_expenses"].copy()
    _require_columns(oa_source, ["product_sub_category", "amount"], "fact_expense OA")
    oa_source["product_sub_category"] = oa_source["product_sub_category"].map(_normalize_product)
    oa_source["amount"] = pd.to_numeric(oa_source["amount"], errors="coerce").fillna(0.0)
    oa_source["in_product_master"] = (
        oa_source["product_sub_category"].map(_product_key).isin(product_master_values)
    )
    excluded_oa_amount = float(oa_source.loc[~oa_source["in_product_master"], "amount"].sum())
    oa_model = oa_source[oa_source["in_product_master"]].copy()
    oa_model["bucket"] = oa_model["product_sub_category"].map(
        lambda value: _bucket_product(value, selected_product_lookup)
    )
    oa = oa_model.groupby("bucket")["amount"].sum()

    has_unmapped = any(
        UNMAPPED_PRODUCT in series.index
        and not np.isclose(float(series.get(UNMAPPED_PRODUCT, 0.0)), 0.0)
        for series in [
            rd_revenue,
            rd_cost,
            electronic_revenue,
            electronic_cost,
            sales_quantity,
            labor,
            material,
            mold,
            tech,
            oa,
        ]
    )
    output_products = products + ([UNMAPPED_PRODUCT] if has_unmapped else [])
    detail = pd.DataFrame({"product_sub_category": output_products})
    detail["project_stage"] = detail["product_sub_category"].map(stage_map)
    detail["mapping_status"] = np.where(
        detail["product_sub_category"].eq(UNMAPPED_PRODUCT), "unmapped", "mapped"
    )

    detail["sales_quantity"] = _series_to_column(detail, sales_quantity)
    detail["rd_related_revenue"] = _series_to_column(detail, rd_revenue)
    detail["rd_related_cost"] = _series_to_column(detail, rd_cost)
    detail["rd_related_gross_profit"] = detail["rd_related_revenue"] - detail["rd_related_cost"]
    detail["rd_related_gross_margin"] = np.divide(
        detail["rd_related_gross_profit"],
        detail["rd_related_revenue"],
        out=np.zeros(len(detail), dtype=float),
        where=~np.isclose(detail["rd_related_revenue"], 0.0),
    )
    detail["electronic_payment_revenue"] = _series_to_column(detail, electronic_revenue)
    detail["electronic_payment_cost"] = _series_to_column(detail, electronic_cost)
    detail["oa_expense"] = _series_to_column(detail, oa)
    detail["material_expense"] = _series_to_column(detail, material)
    detail["mold_expense"] = _series_to_column(detail, mold)
    detail["tech_maintenance_expense"] = _series_to_column(detail, tech)

    labor_total = float(labor.sum())
    income_total = float(electronic_revenue.sum())
    cost_total = float(electronic_cost.sum())
    detail["labor_share"] = (
        _series_to_column(detail, labor) / labor_total if not np.isclose(labor_total, 0.0) else 0.0
    )
    detail["income_share"] = (
        detail["electronic_payment_revenue"] / income_total
        if not np.isclose(income_total, 0.0)
        else 0.0
    )
    detail["cost_share"] = (
        detail["electronic_payment_cost"] / cost_total if not np.isclose(cost_total, 0.0) else 0.0
    )

    expense_bases = sources["expense_bases"].copy()
    _require_columns(expense_bases, ["prim_subj", "amount"], "fact_bus_expense")
    expense_bases["amount"] = pd.to_numeric(expense_bases["amount"], errors="coerce").fillna(0.0)
    base_map = expense_bases.groupby("prim_subj")["amount"].sum().to_dict()
    rd_expense_base = float(base_map.get("研发费用", 0.0))
    management_expense_base = float(base_map.get("管理费用", 0.0))
    sales_expense_base = float(base_map.get("销售费用", 0.0))

    # 保持已经核对通过的分摊池口径：领料、模具、技术维护包含未映射成员，
    # OA 在计算研发费用可分摊池时排除未映射成员。
    selected_material_total = float(material.sum())
    selected_mold_total = float(mold.sum())
    selected_oa_for_rd_pool_total = float(oa.drop(labels=[UNMAPPED_PRODUCT], errors="ignore").sum())
    selected_tech_total = float(tech.sum())
    rd_allocatable_pool = (
        rd_expense_base
        - selected_material_total
        - selected_mold_total
        - selected_oa_for_rd_pool_total
    )
    sales_allocatable_pool = sales_expense_base - selected_tech_total

    detail["allocated_rd_expense"] = detail["labor_share"] * rd_allocatable_pool
    detail["allocated_sales_expense"] = detail["income_share"] * sales_allocatable_pool
    detail["allocated_management_expense"] = detail["cost_share"] * management_expense_base
    expense_components = [
        "oa_expense",
        "material_expense",
        "mold_expense",
        "allocated_rd_expense",
        "tech_maintenance_expense",
        "allocated_sales_expense",
        "allocated_management_expense",
    ]
    detail["total_expense"] = detail[expense_components].sum(axis=1)
    detail["remaining_profit"] = detail["rd_related_gross_profit"] - detail["total_expense"]

    detail["period_start"] = pd.Timestamp(period_start).normalize()
    detail["period_end"] = pd.Timestamp(period_end).normalize()
    detail["rd_expense_base"] = rd_expense_base
    detail["management_expense_base"] = management_expense_base
    detail["sales_expense_base"] = sales_expense_base
    detail["rd_allocatable_pool"] = rd_allocatable_pool
    detail["sales_allocatable_pool"] = sales_allocatable_pool
    detail["source_no"] = detail["product_sub_category"].map(
        lambda product: "RDP-"
        + hashlib.sha1(
            f"{pd.Timestamp(period_start).date()}|{pd.Timestamp(period_end).date()}|{product}".encode(
                "utf-8"
            )
        ).hexdigest()[:20]
    )
    detail["calculated_at"] = pd.Timestamp.now(tz="Asia/Shanghai").tz_localize(None)
    detail = detail[RESULT_COLUMNS].copy()

    metrics = {
        "selected_product_count": len(products),
        "result_row_count": len(detail),
        "revenue_cost_backup_row_count": len(revenue_cost_backup),
        "has_unmapped_product": has_unmapped,
        "excluded_oa_amount": excluded_oa_amount,
        "unmapped_oa_amount": float(oa.get(UNMAPPED_PRODUCT, 0.0)),
        "income_share_total": float(detail["income_share"].sum()),
        "labor_share_total": float(detail["labor_share"].sum()),
        "cost_share_total": float(detail["cost_share"].sum()),
    }
    return {
        "detail": detail,
        "revenue_cost_backup": revenue_cost_backup,
        "metrics": metrics,
    }


@task(name="calculate_rd_project_profitability", log_prints=True)
def calculate_rd_project_profitability_task(
    sources: Mapping[str, pd.DataFrame],
    period_start: pd.Timestamp,
    period_end: pd.Timestamp,
) -> Dict[str, object]:
    """Prefect task 包装：计算研发项目收益明细。"""
    result = calculate_rd_project_profitability(sources, period_start, period_end)
    print(
        "研发项目收益明细计算完成："
        f"products={result['metrics']['selected_product_count']}, "
        f"rows={result['metrics']['result_row_count']}, "
        f"excluded_oa={result['metrics']['excluded_oa_amount']:.2f}"
    )
    return result


def validate_rd_project_profitability(
    result: Mapping[str, object],
    tolerance: float = 0.01,
) -> Dict[str, object]:
    """校验分摊闭合、行级公式和收入成本备查表合计。"""
    if tolerance < 0:
        raise ValueError("tolerance 不能为负数")
    if not {"detail", "revenue_cost_backup", "metrics"}.issubset(result):
        raise ValueError("研发项目收益结果必须包含 detail、revenue_cost_backup 和 metrics")

    detail = result["detail"]
    revenue_cost_backup = result["revenue_cost_backup"]
    metrics = result["metrics"]
    if not isinstance(detail, pd.DataFrame) or detail.empty:
        raise ValueError("研发项目收益明细为空")
    if not isinstance(revenue_cost_backup, pd.DataFrame):
        raise ValueError("收入成本备查数据必须是 DataFrame")
    _require_columns(detail, RESULT_COLUMNS, "研发项目收益结果")
    _require_columns(
        revenue_cost_backup,
        REVENUE_COST_BACKUP_COLUMNS,
        "研发项目收益收入成本备查数据",
    )

    component_sum = detail[
        [
            "oa_expense",
            "material_expense",
            "mold_expense",
            "allocated_rd_expense",
            "tech_maintenance_expense",
            "allocated_sales_expense",
            "allocated_management_expense",
        ]
    ].sum(axis=1)
    residuals = {
        "expense_component_max_residual": float(
            (detail["total_expense"] - component_sum).abs().max()
        ),
        "gross_profit_max_residual": float(
            (
                detail["rd_related_gross_profit"]
                - (detail["rd_related_revenue"] - detail["rd_related_cost"])
            )
            .abs()
            .max()
        ),
        "remaining_profit_max_residual": float(
            (
                detail["remaining_profit"]
                - (detail["rd_related_gross_profit"] - detail["total_expense"])
            )
            .abs()
            .max()
        ),
        "rd_allocation_residual": float(
            detail["allocated_rd_expense"].sum() - detail["rd_allocatable_pool"].iloc[0]
        ),
        "sales_allocation_residual": float(
            detail["allocated_sales_expense"].sum() - detail["sales_allocatable_pool"].iloc[0]
        ),
        "management_allocation_residual": float(
            detail["allocated_management_expense"].sum() - detail["management_expense_base"].iloc[0]
        ),
        "backup_revenue_residual": float(
            revenue_cost_backup["revenue"].sum() - detail["rd_related_revenue"].sum()
        ),
        "backup_cost_residual": float(
            revenue_cost_backup["cost"].sum() - detail["rd_related_cost"].sum()
        ),
    }
    failed_residuals = {name: value for name, value in residuals.items() if abs(value) > tolerance}
    if failed_residuals:
        details = ", ".join(f"{name}={value:.6f}" for name, value in failed_residuals.items())
        raise ValueError(f"研发项目收益公式或分摊未闭合: {details}")

    for share_name in ["income_share_total", "labor_share_total", "cost_share_total"]:
        share_total = float(metrics[share_name])
        if not np.isclose(share_total, 0.0) and abs(share_total - 1.0) > tolerance:
            raise ValueError(f"{share_name} 未闭合到 1: {share_total:.6f}")

    return {
        **residuals,
        "row_count": len(detail),
        "rd_related_revenue_total": float(detail["rd_related_revenue"].sum()),
        "rd_related_cost_total": float(detail["rd_related_cost"].sum()),
        "rd_related_gross_profit_total": float(detail["rd_related_gross_profit"].sum()),
        "total_expense": float(detail["total_expense"].sum()),
        "remaining_profit_total": float(detail["remaining_profit"].sum()),
        "revenue_cost_backup_row_count": len(revenue_cost_backup),
        "revenue_cost_backup_revenue_total": float(revenue_cost_backup["revenue"].sum()),
        "revenue_cost_backup_cost_total": float(revenue_cost_backup["cost"].sum()),
        "excluded_oa_amount": float(metrics["excluded_oa_amount"]),
    }


@task(name="validate_rd_project_profitability", log_prints=True)
def validate_rd_project_profitability_task(
    result: Mapping[str, object],
    tolerance: float = 0.01,
) -> Dict[str, object]:
    """Prefect task 包装：校验研发项目收益结果。"""
    validation = validate_rd_project_profitability(result, tolerance)
    print(
        "研发项目收益校验通过："
        f"rows={validation['row_count']}, "
        f"expense={validation['total_expense']:.2f}, "
        f"backup_rows={validation['revenue_cost_backup_row_count']}"
    )
    return validation


def _format_excel_workbook(
    writer,
    detail_sheet: str,
    summary_sheet: str,
    rules_sheet: str,
    backup_sheet: str,
) -> None:
    """为前端下载的 Excel 添加基础可读性格式。"""
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = writer.book
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    percentage_headers = {
        "研发相关毛利率",
        "收入占比",
        "工时占比",
        "成本占比",
    }
    amount_headers = {
        "本期研发相关收入",
        "本期研发相关成本",
        "研发相关毛利",
        "费用总额",
        "剩余收益",
        "OA费用",
        "领料费用",
        "模具费用",
        "研发费用分摊",
        "技术维护费",
        "销售费用分摊",
        "管理费用分摊",
        "研发费用基数",
        "管理费用基数",
        "销售费用基数",
        "收入",
        "成本",
        "毛利",
    }

    for sheet_name in [detail_sheet, summary_sheet, rules_sheet, backup_sheet]:
        worksheet = workbook[sheet_name]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in worksheet.columns:
            values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
            width = min(max(max(len(value) for value in values) + 2, 10), 34)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width

    detail_worksheet = workbook[detail_sheet]
    detail_headers = {cell.value: cell.column for cell in detail_worksheet[1]}
    for header in percentage_headers:
        column = detail_headers.get(header)
        if column:
            for row in range(2, detail_worksheet.max_row + 1):
                detail_worksheet.cell(row=row, column=column).number_format = "0.00%"
    for sheet_name in [detail_sheet, backup_sheet]:
        worksheet = workbook[sheet_name]
        headers = {cell.value: cell.column for cell in worksheet[1]}
        for header in amount_headers:
            column = headers.get(header)
            if column:
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(row=row, column=column).number_format = "#,##0.00"

    backup_worksheet = workbook[backup_sheet]
    backup_headers = {cell.value: cell.column for cell in backup_worksheet[1]}
    for header in ["开始日期", "结束日期"]:
        column = backup_headers.get(header)
        if column:
            for row in range(2, backup_worksheet.max_row + 1):
                backup_worksheet.cell(row=row, column=column).number_format = "yyyy-mm-dd"


def _safe_excel_text(value: object) -> object:
    """阻止数据库文本被 Excel 解释为公式。"""
    if not isinstance(value, str):
        return value
    return f"'{value}" if value.startswith(("=", "+", "-", "@")) else value


def export_rd_project_profitability_excel(
    result: Mapping[str, object],
    validation: Mapping[str, object],
    output_dir: Optional[str] = None,
    download_base_url: Optional[str] = None,
) -> Dict[str, object]:
    """生成可供前端下载的研发项目收益 Excel，并返回文件元数据。"""
    detail = result.get("detail")
    revenue_cost_backup = result.get("revenue_cost_backup")
    if not isinstance(detail, pd.DataFrame) or detail.empty:
        raise ValueError("没有可导出的研发项目收益明细")
    if not isinstance(revenue_cost_backup, pd.DataFrame):
        raise ValueError("没有可导出的收入成本备查数据")
    _require_columns(detail, RESULT_COLUMNS, "研发项目收益导出数据")
    _require_columns(
        revenue_cost_backup,
        REVENUE_COST_BACKUP_COLUMNS,
        "研发项目收益收入成本备查数据",
    )

    configured_output_dir = output_dir or os.environ.get(
        "RD_PROJECT_OUTPUT_DIR", DEFAULT_OUTPUT_DIR
    )
    output_path_root = Path(configured_output_dir).expanduser()
    if not output_path_root.is_absolute():
        raise ValueError("output_dir 必须是绝对路径")
    output_path_root.mkdir(parents=True, exist_ok=True)

    period_start = pd.Timestamp(detail["period_start"].iloc[0])
    period_end = pd.Timestamp(detail["period_end"].iloc[0])
    generated_at = pd.Timestamp.now(tz="Asia/Shanghai")
    filename = (
        "研发项目收益分析_"
        f"{period_start.strftime('%Y%m%d')}_{period_end.strftime('%Y%m%d')}_"
        f"{generated_at.strftime('%Y%m%d%H%M%S%f')}.xlsx"
    )
    output_path = output_path_root / filename

    export_columns = {
        "product_sub_category": "产品细类",
        "project_stage": "项目所属阶段",
        "mapping_status": "映射状态",
        "sales_quantity": "本期销量(个)",
        "rd_related_revenue": "本期研发相关收入",
        "rd_related_cost": "本期研发相关成本",
        "rd_related_gross_profit": "研发相关毛利",
        "rd_related_gross_margin": "研发相关毛利率",
        "total_expense": "费用总额",
        "remaining_profit": "剩余收益",
        "oa_expense": "OA费用",
        "material_expense": "领料费用",
        "mold_expense": "模具费用",
        "allocated_rd_expense": "研发费用分摊",
        "tech_maintenance_expense": "技术维护费",
        "allocated_sales_expense": "销售费用分摊",
        "allocated_management_expense": "管理费用分摊",
        "income_share": "收入占比",
        "labor_share": "工时占比",
        "cost_share": "成本占比",
        "rd_expense_base": "研发费用基数",
        "management_expense_base": "管理费用基数",
        "sales_expense_base": "销售费用基数",
    }
    detail_export = detail[list(export_columns)].rename(columns=export_columns)
    detail_export["映射状态"] = detail_export["映射状态"].replace({"mapped": "已映射", "unmapped": "未映射"})
    for text_column in ["产品细类", "项目所属阶段", "映射状态"]:
        detail_export[text_column] = detail_export[text_column].map(_safe_excel_text)

    backup_columns = {
        "period_start": "开始日期",
        "period_end": "结束日期",
        "product_sub_category": "产品细类",
        "mapped_product_sub_category": "主数据产品细类",
        "mapping_status": "映射状态",
        "material_code": "物料编码",
        "revenue_category": "收入大类",
        "primary_org": "一级组织",
        "secondary_org": "二级组织",
        "revenue": "收入",
        "cost": "成本",
        "gross_profit": "毛利",
    }
    backup_export = revenue_cost_backup[list(backup_columns)].rename(columns=backup_columns)
    backup_export["映射状态"] = backup_export["映射状态"].replace({"mapped": "已映射", "unmapped": "未映射"})
    for text_column in [
        "产品细类",
        "主数据产品细类",
        "映射状态",
        "物料编码",
        "收入大类",
        "一级组织",
        "二级组织",
    ]:
        backup_export[text_column] = backup_export[text_column].map(_safe_excel_text)

    summary_rows = [
        ("开始日期", period_start.strftime("%Y-%m-%d")),
        ("结束日期", period_end.strftime("%Y-%m-%d")),
        ("生成时间", generated_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("项目行数", int(validation["row_count"])),
        ("研发相关收入", float(validation["rd_related_revenue_total"])),
        ("研发相关成本", float(validation["rd_related_cost_total"])),
        ("研发相关毛利", float(validation["rd_related_gross_profit_total"])),
        ("费用总额", float(validation["total_expense"])),
        ("剩余收益", float(validation["remaining_profit_total"])),
        ("收入成本备查行数", int(validation["revenue_cost_backup_row_count"])),
        ("备查收入合计", float(validation["revenue_cost_backup_revenue_total"])),
        ("备查成本合计", float(validation["revenue_cost_backup_cost_total"])),
        ("产品主数据未承接OA费用", float(validation["excluded_oa_amount"])),
    ]
    summary_export = pd.DataFrame(summary_rows, columns=["指标", "值"])
    rules_export = pd.DataFrame(
        [
            ("研发相关收入", "二级组织为国际业务中心的电子支付收入和技术服务收入"),
            ("研发相关成本", "二级组织为国际业务中心的电子支付成本和技术服务成本"),
            ("研发费用分摊", "国际业务研发费用基数扣除选中产品OA/领料/模具后，按研发工时占比分摊"),
            ("销售费用分摊", "国际业务销售费用基数扣除技术维护费后，按电子支付收入占比分摊"),
            ("管理费用分摊", "国际业务管理费用基数按电子支付成本占比分摊"),
            ("费用总额", "OA + 领料 + 模具 + 研发分摊 + 技术维护 + 销售分摊 + 管理分摊"),
            ("研发相关毛利率", "研发相关毛利除以研发相关收入"),
            ("剩余收益", "研发相关毛利 - 费用总额"),
            ("收入成本备查", "仅包含国际业务中心的电子支付和技术服务收入成本明细"),
            ("未映射产品", "表示产品主数据或研发项目阶段未完整映射的数据"),
        ],
        columns=["指标", "计算口径"],
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        detail_export.to_excel(writer, sheet_name="研发项目收益", index=False)
        summary_export.to_excel(writer, sheet_name="汇总与校验", index=False)
        rules_export.to_excel(writer, sheet_name="计算口径", index=False)
        backup_export.to_excel(writer, sheet_name="收入成本备查", index=False)
        _format_excel_workbook(
            writer,
            "研发项目收益",
            "汇总与校验",
            "计算口径",
            "收入成本备查",
        )

    configured_base_url = download_base_url or os.environ.get("RD_PROJECT_DOWNLOAD_BASE_URL")
    download_url = None
    if configured_base_url:
        download_url = f"{configured_base_url.rstrip('/')}/{quote(filename)}"
    return {
        "file_name": filename,
        "output_path": str(output_path),
        "download_url": download_url,
        "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "size_bytes": output_path.stat().st_size,
        "row_count": len(detail_export),
        "backup_row_count": len(backup_export),
        "sheet_names": ["研发项目收益", "汇总与校验", "计算口径", "收入成本备查"],
    }


@task(name="export_rd_project_profitability_excel", log_prints=True)
def export_rd_project_profitability_excel_task(
    result: Mapping[str, object],
    validation: Mapping[str, object],
    output_dir: Optional[str] = None,
    download_base_url: Optional[str] = None,
) -> Dict[str, object]:
    """Prefect task 包装：生成研发项目收益 Excel。"""
    report = export_rd_project_profitability_excel(
        result=result,
        validation=validation,
        output_dir=output_dir,
        download_base_url=download_base_url,
    )
    print(
        f"研发项目收益 Excel 已生成：{report['output_path']}，"
        f"rows={report['row_count']}, size={report['size_bytes']}"
    )
    return report


def _validate_target_table(target_table: str) -> str:
    if not isinstance(target_table, str) or not TARGET_TABLE_PATTERN.fullmatch(target_table):
        raise ValueError("target_table 只能包含小写字母、数字和下划线，且不能以数字开头")
    return target_table


@task(name="replace_rd_project_profitability_snapshot", log_prints=True)
def replace_rd_project_profitability_snapshot_task(
    result: Mapping[str, object],
    target_table: str = "fact_rd_project_profitability",
) -> Dict[str, object]:
    """事务替换指定期间的研发项目收益快照。"""
    table_name = _validate_target_table(target_table)
    detail = result.get("detail")
    if not isinstance(detail, pd.DataFrame) or detail.empty:
        raise ValueError("没有可写入的研发项目收益明细")
    _require_columns(detail, RESULT_COLUMNS, "研发项目收益写入数据")

    insert_columns = RESULT_COLUMNS
    placeholders = ", ".join(["%s"] * len(insert_columns))
    rows = []
    for record in detail[insert_columns].itertuples(index=False, name=None):
        rows.append(
            tuple(
                None
                if pd.isna(value)
                else value.to_pydatetime()
                if isinstance(value, pd.Timestamp)
                else value.item()
                if isinstance(value, np.generic)
                else value
                for value in record
            )
        )

    period_start = pd.Timestamp(detail["period_start"].iloc[0]).date()
    period_end = pd.Timestamp(detail["period_end"].iloc[0]).date()
    conn, cur = connect_to_db()
    try:
        cur.execute(
            "SELECT pg_advisory_xact_lock(hashtext(%s))",
            (f"{table_name}:{period_start}:{period_end}",),
        )
        cur.execute(
            f"""
            CREATE TABLE IF NOT EXISTS {table_name} (
                period_start DATE NOT NULL,
                period_end DATE NOT NULL,
                product_sub_category TEXT NOT NULL,
                project_stage TEXT,
                mapping_status TEXT NOT NULL,
                sales_quantity NUMERIC,
                rd_related_revenue NUMERIC,
                rd_related_cost NUMERIC,
                rd_related_gross_profit NUMERIC,
                rd_related_gross_margin NUMERIC,
                total_expense NUMERIC,
                remaining_profit NUMERIC,
                oa_expense NUMERIC,
                material_expense NUMERIC,
                mold_expense NUMERIC,
                allocated_rd_expense NUMERIC,
                tech_maintenance_expense NUMERIC,
                allocated_sales_expense NUMERIC,
                allocated_management_expense NUMERIC,
                income_share NUMERIC,
                labor_share NUMERIC,
                cost_share NUMERIC,
                rd_expense_base NUMERIC,
                management_expense_base NUMERIC,
                sales_expense_base NUMERIC,
                rd_allocatable_pool NUMERIC,
                sales_allocatable_pool NUMERIC,
                source_no TEXT NOT NULL,
                calculated_at TIMESTAMP WITHOUT TIME ZONE NOT NULL,
                PRIMARY KEY (period_start, period_end, product_sub_category)
            )
            """
        )
        cur.execute(
            f"DELETE FROM {table_name} WHERE period_start = %s AND period_end = %s",
            (period_start, period_end),
        )
        deleted_rows = cur.rowcount
        cur.executemany(
            f"INSERT INTO {table_name} ({', '.join(insert_columns)}) VALUES ({placeholders})",
            rows,
        )
        inserted_rows = cur.rowcount
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()

    print(
        f"研发项目收益快照写入完成：table={table_name}, deleted={deleted_rows}, "
        f"inserted={inserted_rows}, period={period_start}..{period_end}"
    )
    return {
        "target_table": table_name,
        "period_start": period_start,
        "period_end": period_end,
        "deleted_rows": deleted_rows,
        "inserted_rows": inserted_rows,
    }
