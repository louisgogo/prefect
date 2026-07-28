import tempfile
import unittest
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from modules.rd_project_profitability.tasks.rd_project_profitability_tasks import (
    UNMAPPED_PRODUCT,
    calculate_rd_project_profitability,
    export_rd_project_profitability_excel,
    get_default_rd_project_period,
    resolve_rd_project_period,
    validate_rd_project_profitability,
)


def _sources():
    return {
        "project_stages": pd.DataFrame(
            [
                {"product_sub_category": "A", "proj_status": "量产项目"},
                {"product_sub_category": "a ", "proj_status": "量产项目"},
                {"product_sub_category": "B", "proj_status": "开发转产项目"},
            ]
        ),
        "material_master": pd.DataFrame(
            [
                {
                    "encoding": "m1",
                    "name": "A产品",
                    "material_group": "产成品",
                    "material_major_category": None,
                },
                {
                    "encoding": "m2",
                    "name": "B产品",
                    "material_group": "产成品",
                    "material_major_category": None,
                },
                {
                    "encoding": "m3",
                    "name": "C产品",
                    "material_group": "产成品",
                    "material_major_category": None,
                },
            ]
        ),
        "product_overrides": pd.DataFrame(columns=["encoding", "product_sub_category"]),
        "revenue": pd.DataFrame(
            [
                {
                    "mat_code": "m1",
                    "inc_major_cat": "电子支付",
                    "prim_org": "硬件事业群",
                    "sec_org": "其他中心",
                    "sales_quantity": 10,
                    "revenue": 100,
                    "cost": 60,
                },
                {
                    "mat_code": "m1",
                    "inc_major_cat": "技术服务",
                    "prim_org": "渠道事业群",
                    "sec_org": "国内市场中心",
                    "sales_quantity": 0,
                    "revenue": 50,
                    "cost": 20,
                },
                {
                    "mat_code": "m1",
                    "inc_major_cat": "电子支付",
                    "prim_org": "硬件事业群",
                    "sec_org": "国际业务中心",
                    "sales_quantity": 8,
                    "revenue": 80,
                    "cost": 40,
                },
                {
                    "mat_code": "m1",
                    "inc_major_cat": "技术服务",
                    "prim_org": "渠道事业群",
                    "sec_org": "国际业务中心",
                    "sales_quantity": 0,
                    "revenue": 40,
                    "cost": 10,
                },
                {
                    "mat_code": "m2",
                    "inc_major_cat": "电子支付",
                    "prim_org": "硬件事业群",
                    "sec_org": "其他中心",
                    "sales_quantity": 20,
                    "revenue": 300,
                    "cost": 90,
                },
                {
                    "mat_code": "missing",
                    "inc_major_cat": "电子支付",
                    "prim_org": "硬件事业群",
                    "sec_org": "其他中心",
                    "sales_quantity": 5,
                    "revenue": 100,
                    "cost": 50,
                },
            ]
        ),
        "oa_expenses": pd.DataFrame(
            [
                {"product_sub_category": "A", "amount": 40},
                {"product_sub_category": "C", "amount": 50},
                {"product_sub_category": "D", "amount": 60},
            ]
        ),
        "labor_hours": pd.DataFrame(
            [
                {"product_sub_category": "A", "amount": 1},
                {"product_sub_category": "B", "amount": 3},
                {"product_sub_category": "C", "amount": 1},
            ]
        ),
        "material_expenses": pd.DataFrame(
            [
                {"product_sub_category": "A", "amount": 10},
                {"product_sub_category": "C", "amount": 5},
            ]
        ),
        "mold_expenses": pd.DataFrame([{"product_sub_category": "B", "amount": 20}]),
        "tech_maintenance_expenses": pd.DataFrame([{"product_sub_category": "B", "amount": 30}]),
        "expense_bases": pd.DataFrame(
            [
                {"prim_subj": "研发费用", "amount": 200},
                {"prim_subj": "销售费用", "amount": 130},
                {"prim_subj": "管理费用", "amount": 80},
            ]
        ),
    }


class RdProjectProfitabilityTests(unittest.TestCase):
    def test_default_period_is_year_to_most_recent_completed_month(self):
        start, end = get_default_rd_project_period(date(2026, 7, 28))

        self.assertEqual(start, pd.Timestamp("2026-01-01"))
        self.assertEqual(end, pd.Timestamp("2026-06-30"))

    def test_explicit_period_requires_both_dates_and_order(self):
        with self.assertRaisesRegex(ValueError, "同时提供"):
            resolve_rd_project_period("2026-01-01", None)
        with self.assertRaisesRegex(ValueError, "不能晚于"):
            resolve_rd_project_period("2026-07-01", "2026-06-30")

    def test_calculation_closes_allocations_and_exposes_unmapped_bucket(self):
        result = calculate_rd_project_profitability(
            _sources(), pd.Timestamp("2026-01-01"), pd.Timestamp("2026-06-30")
        )
        validation = validate_rd_project_profitability(result)
        detail = result["detail"].set_index("product_sub_category")

        self.assertEqual(result["metrics"]["selected_product_count"], 2)
        self.assertIn(UNMAPPED_PRODUCT, detail.index)
        self.assertAlmostEqual(detail["income_share"].sum(), 1.0)
        self.assertAlmostEqual(detail["labor_share"].sum(), 1.0)
        self.assertAlmostEqual(detail["cost_share"].sum(), 1.0)
        self.assertAlmostEqual(detail["rd_allocatable_pool"].iloc[0], 125.0)
        self.assertAlmostEqual(detail["sales_allocatable_pool"].iloc[0], 100.0)
        self.assertAlmostEqual(detail["allocated_rd_expense"].sum(), 125.0)
        self.assertAlmostEqual(detail["allocated_sales_expense"].sum(), 100.0)
        self.assertAlmostEqual(detail["allocated_management_expense"].sum(), 80.0)
        self.assertAlmostEqual(validation["excluded_oa_amount"], 60.0)
        self.assertAlmostEqual(validation["total_expense"], 460.0)
        self.assertAlmostEqual(validation["power_bi_grand_total_expense"], 410.0)
        self.assertAlmostEqual(validation["power_bi_non_additive_expense_gap"], 50.0)
        self.assertAlmostEqual(validation["remaining_profit_total"], -610.0)
        self.assertAlmostEqual(validation["power_bi_remaining_profit_total"], -120.0)
        self.assertAlmostEqual(validation["remaining_profit_gap_total"], -490.0)
        self.assertAlmostEqual(detail.loc["A", "rd_related_revenue"], 120.0)
        self.assertAlmostEqual(detail.loc["A", "electronic_payment_revenue"], 180.0)
        self.assertAlmostEqual(detail.loc["A", "rd_related_cost"], 130.0)
        self.assertAlmostEqual(detail.loc["A", "rd_related_gross_margin"], -10 / 120)
        self.assertAlmostEqual(detail.loc["A", "power_bi_gross_margin"], 80 / 180)
        self.assertAlmostEqual(detail.loc["B", "rd_related_revenue"], 0.0)
        self.assertAlmostEqual(detail.loc[UNMAPPED_PRODUCT, "rd_related_revenue"], 0.0)

    def test_excel_export_returns_frontend_file_metadata(self):
        result = calculate_rd_project_profitability(
            _sources(), pd.Timestamp("2026-01-01"), pd.Timestamp("2026-06-30")
        )
        validation = validate_rd_project_profitability(result)

        with tempfile.TemporaryDirectory() as output_dir:
            report = export_rd_project_profitability_excel(
                result=result,
                validation=validation,
                output_dir=output_dir,
                download_base_url="https://reports.example/rd",
            )
            output_path = Path(report["output_path"])
            workbook = load_workbook(output_path, read_only=True, data_only=True)

            self.assertTrue(output_path.exists())
            self.assertEqual(report["row_count"], 3)
            self.assertEqual(workbook.sheetnames, ["研发项目收益", "汇总与校验", "计算口径"])
            self.assertTrue(report["download_url"].startswith("https://reports.example/rd/"))
            self.assertGreater(report["size_bytes"], 0)


if __name__ == "__main__":
    unittest.main()
